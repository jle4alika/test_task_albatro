import os
import json
import openpyxl

from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.copier import WorksheetCopy
import shutil
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Border, Side

def save_data(filename: str, links: list):
    #  Создаём пустой файл и сохраняем его
    wb = openpyxl.Workbook()
    wb.save(filename)

    #  Открываем созданный нами файл
    wb = openpyxl.load_workbook(filename, data_only=True)
    #  Называем лист
    wb.title = 'dsn-shop.ru'

    sheet = wb.active
    #  Проставляем заголовки в таблице и выравнивание, обрамление
    sheet['A1'].value = 'Категория'
    sheet['A1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['A1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['B1'].value = 'Наименование'
    sheet['B1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['B1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['C1'].value = 'Цена'
    sheet['C1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['C1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['D1'].value = 'Доступность'
    sheet['D1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['D1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['E1'].value = 'Ссылка на товар'
    sheet['E1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['E1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['F1'].value = 'Ссылка на главное изображение'
    sheet['F1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['F1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['G1'].value = 'Ссылки на все изображения'
    sheet['G1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['G1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['H1'].value = 'Характеристики'
    sheet['H1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['H1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))

    sheet['I1'].value = 'Описание'
    sheet['I1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['I1'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'),
                                      left=Side(border_style='thin', color='FF000000'))


     # Получаем данные уже спарщеных товаров
    with open('products_about.json', 'r') as file:
        products_about_dict: dict = json.load(file)

    rows = 2
    for links in links:
        for i in range(len(links)):
            #  Заполняем таблицу данными
            sheet[f'A{rows}'].value = products_about_dict[links[i]]['Категория']
            sheet[f'A{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'B{rows}'].value = products_about_dict[links[i]]['Наименование']
            sheet[f'B{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'C{rows}'].value = products_about_dict[links[i]]['Цена']
            sheet[f'C{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'D{rows}'].value = products_about_dict[links[i]]['Доступность']
            sheet[f'D{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'E{rows}'].value = products_about_dict[links[i]]['Ссылка на товар']
            sheet[f'E{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'F{rows}'].value = products_about_dict[links[i]]['Ссылка на главное изображение']
            sheet[f'F{rows}'].alignment = Alignment(horizontal="center", vertical='center')

            sheet[f'G{rows}'].value = products_about_dict[links[i]]['Ссылки на все изображения']
            sheet[f'G{rows}'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

            sheet[f'H{rows}'].value = products_about_dict[links[i]]['Характеристики']

            sheet[f'I{rows}'].value = products_about_dict[links[i]]['Описание']
            sheet[f'I{rows}'].alignment = Alignment(wrap_text=True)

            # Выставляем размеры строк и столбцов
            sheet.row_dimensions[rows].height = 30

            sheet.column_dimensions['A'].width = 25
            sheet.column_dimensions['B'].width = 50
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 40
            sheet.column_dimensions['E'].width = 100
            sheet.column_dimensions['F'].width = 200
            sheet.column_dimensions['G'].width = 200
            sheet.column_dimensions['H'].width = 25
            sheet.column_dimensions['I'].width = 100

            #  Устанавливаем обрамление
            sheet[f'A{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'B{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'C{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'D{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'E{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'F{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'G{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'H{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            sheet[f'I{rows}'].border = Border(top=Side(border_style='thin', color='FF000000'),
                                                 right=Side(border_style='thin', color='FF000000'),
                                                 bottom=Side(border_style='thin', color='FF000000'),
                                                 left=Side(border_style='thin', color='FF000000'))

            rows += 1

    #  Сохраняем изменения в файле
    wb.save(filename)

